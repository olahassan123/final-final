"""
Run once to import MeDay_Treatments_Data_finalalmost.xlsx into chatbot.db.
Re-run whenever the Excel is updated — it drops and reimports all data tables.

Usage:
    python import_meday_data.py
"""
import openpyxl
from chatbot_db import get_chatbot_db, init_chatbot_db
from chatbot_config import BASE_DIR, EXCEL_FILENAME


def _v(cell):
    """Return cell value as clean string, or None if empty."""
    val = cell
    if val is None:
        return None
    s = str(val).strip()
    return None if s == "" or s.lower() == "none" or s.lower() == "nan" else s


def _int(val):
    try:
        return int(val) if val is not None else None
    except (ValueError, TypeError):
        return None


def import_all():
    path = BASE_DIR / EXCEL_FILENAME
    if not path.exists():
        raise FileNotFoundError(f"Excel not found: {path}")

    print(f"Reading {path} …")
    wb = openpyxl.load_workbook(str(path))

    init_chatbot_db()
    conn = get_chatbot_db()

    # ── Drop and reimport all data tables (not sessions) ─────────────
    for tbl in ("cb_scoring", "cb_questions", "cb_faq", "cb_forward_topics", "cb_treatments", "cb_categories"):
        conn.execute(f"DELETE FROM {tbl}")

    # ── Categories ────────────────────────────────────────────────────
    ws = wb["Categories"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    count = 0
    for row in rows[1:]:
        d = dict(zip(headers, row))
        cat_id = _v(d.get("category_id"))
        if not cat_id:
            continue
        conn.execute(
            "INSERT OR REPLACE INTO cb_categories (category_id, category_name, has_recommendation, recommendation_intro) VALUES (?, ?, ?, ?)",
            (cat_id, _v(d.get("category_name")), 1 if _v(d.get("has_recommendation")) == "1" else 0, _v(d.get("recommendation_intro"))),
        )
        count += 1
    print(f"  Categories: {count}")

    # ── Treatments ────────────────────────────────────────────────────
    ws = wb["Treatments"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    count = 0
    for row in rows[1:]:
        d = dict(zip(headers, row))
        tid = _v(d.get("treatment_id"))
        if not tid:
            continue
        conn.execute("""
            INSERT OR REPLACE INTO cb_treatments
            (treatment_id, category_id, subgroup, treatment_name, short_description,
             good_for, technique_or_equipment, duration_min, duration_notes,
             aliases, canonical_id, source)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            tid,
            _v(d.get("category_id")),
            _v(d.get("subgroup")),
            _v(d.get("treatment_name")),
            _v(d.get("short_description")),
            _v(d.get("good_for")),
            _v(d.get("technique_or_equipment")),
            _int(d.get("duration_min")),
            _v(d.get("duration_notes")),
            _v(d.get("aliases")),
            _v(d.get("canonical_id")),
            _v(d.get("source")),
        ))
        count += 1
    print(f"  Treatments: {count}")

    # ── Recommendation_Questions ──────────────────────────────────────
    ws = wb["Recommendation_Questions"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    count = 0
    for row in rows[1:]:
        d = dict(zip(headers, row))
        if not _v(d.get("category_id")) or not _v(d.get("question_id")):
            continue
        conn.execute("""
            INSERT INTO cb_questions
            (category_id, question_id, question_order, question_text,
             option_value, option_label, option_order, terminal_treatment_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            _v(d.get("category_id")),
            _v(d.get("question_id")),
            _int(d.get("question_order")) or 0,
            _v(d.get("question_text")),
            _v(d.get("option_value")),
            _v(d.get("option_label")),
            _int(d.get("option_order")) or 0,
            _v(d.get("terminal_treatment_id")),
        ))
        count += 1
    print(f"  Questions (option rows): {count}")

    # ── Recommendation_Scoring ────────────────────────────────────────
    ws = wb["Recommendation_Scoring"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    count = 0
    for row in rows[1:]:
        d = dict(zip(headers, row))
        if not _v(d.get("category_id")) or not _v(d.get("treatment_id")):
            continue
        conn.execute("""
            INSERT INTO cb_scoring (category_id, question_id, option_value, treatment_id, score_weight)
            VALUES (?, ?, ?, ?, ?)
        """, (
            _v(d.get("category_id")),
            _v(d.get("question_id")),
            _v(d.get("option_value")),
            _v(d.get("treatment_id")),
            _int(d.get("score_weight")) or 0,
        ))
        count += 1
    print(f"  Scoring rows: {count}")

    # ── FAQ ───────────────────────────────────────────────────────────
    ws = wb["FAQ"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    count = 0
    for row in rows[1:]:
        d = dict(zip(headers, row))
        fid = _v(d.get("faq_id"))
        if not fid:
            continue
        conn.execute("""
            INSERT OR REPLACE INTO cb_faq (faq_id, category_id, canonical_question, answer, example_phrasings)
            VALUES (?, ?, ?, ?, ?)
        """, (
            fid,
            _v(d.get("category_id")),
            _v(d.get("canonical_question")),
            _v(d.get("answer")),
            _v(d.get("example_phrasings")),
        ))
        count += 1
    print(f"  FAQ entries: {count}")

    # ── Forward_Topics ────────────────────────────────────────────────
    ws = wb["Forward_Topics"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    count = 0
    for row in rows[1:]:
        d = dict(zip(headers, row))
        tid = _v(d.get("topic_id"))
        if not tid:
            continue
        conn.execute("""
            INSERT OR REPLACE INTO cb_forward_topics (topic_id, topic_name, reason, example_phrasings, phone_number)
            VALUES (?, ?, ?, ?, ?)
        """, (
            tid,
            _v(d.get("topic_name")),
            _v(d.get("reason")),
            _v(d.get("example_phrasings")),
            _v(d.get("phone_number")),
        ))
        count += 1
    print(f"  Forward topics: {count}")

    conn.commit()
    conn.close()
    print("Import complete - OK")


if __name__ == "__main__":
    import_all()
