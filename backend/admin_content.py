"""
Admin editor for the chatbot's knowledge base — the cb_* tables in chatbot.db.

The chatbot answers strictly from these tables (see chatbot_db.py), so an edit
here changes what the bot says on the very next message: no import script, no
restart. That is the whole point of this module — the old flow required someone
to hand-edit MeDay_Treatments_Data_finalalmost.xlsx and run import_meday_data.py.

Wired into main.py via build_router(), which receives the admin dependency and
the audit logger as arguments so this module never imports main (circular).

Everything is driven by SECTIONS below: the CRUD routes, the form metadata the
admin page renders, and the Excel export/import columns all read from one spec,
so a field is added in exactly one place.
"""
import io
import json
import re
import secrets
import sqlite3
from collections import Counter
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Optional, List, Dict, Any

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse

from chatbot_db import get_chatbot_db, CHATBOT_DB_PATH

BASE_DIR = Path(__file__).parent
BACKUP_DIR = BASE_DIR / "backups"
STAGING_DIR = BASE_DIR / ".import_staging"
STAGING_TTL_MINUTES = 30
KEEP_BACKUPS = 10
MAX_UPLOAD_BYTES = 5 * 1024 * 1024

# The pseudo-category the FAQ table uses for questions that belong to no single
# category. It is a real value in cb_faq.category_id, not a bug — keep it valid.
GENERAL_CATEGORY = "GENERAL"
GENERAL_CATEGORY_LABEL = "כללי"

ID_COLUMN_LABEL = "מזהה"
CATEGORY_NAME_LABEL = "קטגוריה"
CATEGORY_CODE_LABEL = "קוד קטגוריה"


# -- Section spec -------------------------------------------------------------
# type: text | textarea | select | combo | list | bool
#   list  -> stored comma-separated in one column, edited as chips in the UI
#   bool  -> stored as INTEGER 0/1
#   combo -> free text with suggestions from existing values
# Columns NOT listed here are never written by this module. That is deliberate:
# duration_min / duration_notes stay untouched because the bot is forbidden from
# stating durations (SYSTEM_PROMPT rule 5), and canonical_id / source are
# bookkeeping set once at creation.

SECTIONS: Dict[str, Dict[str, Any]] = {
    "treatments": {
        "table": "cb_treatments",
        "pk": "treatment_id",
        "label": "טיפולים",
        "description": "כל טיפול שהצ׳אטבוט יכול לספר עליו. כל טיפול שייך לקטגוריה אחת.",
        "order_by": "category_id, treatment_id",
        "sheet": "טיפולים",
        "name_field": "treatment_name",
        "fields": [
            {"key": "category_id", "label": CATEGORY_NAME_LABEL, "type": "select", "required": True,
             "help": "כל טיפול חייב להשתייך לקטגוריה אחת קיימת."},
            {"key": "subgroup", "label": "תת-קטגוריה", "type": "combo", "required": False,
             "help": "קיבוץ פנימי בתוך הקטגוריה, למשל מניקור או פדיקור. אפשר להשאיר ריק."},
            {"key": "treatment_name", "label": "שם הטיפול", "type": "text", "required": True},
            {"key": "short_description", "label": "תיאור קצר", "type": "textarea", "required": False,
             "help": "במשפט או שניים — מה הטיפול כולל."},
            {"key": "good_for", "label": "למי הטיפול מתאים", "type": "textarea", "required": False},
            {"key": "technique_or_equipment", "label": "טכניקה או מכשור", "type": "text", "required": False},
            {"key": "what_to_expect", "label": "מה קורה במהלך הטיפול", "type": "textarea", "required": False},
            {"key": "preparation", "label": "הכנה לפני הטיפול", "type": "textarea", "required": False},
            {"key": "aftercare", "label": "הנחיות אחרי הטיפול", "type": "textarea", "required": False},
            {"key": "downtime", "label": "זמן החלמה", "type": "text", "required": False},
            {"key": "pain_level", "label": "רמת כאב", "type": "text", "required": False},
            {"key": "sessions_recommended", "label": "מספר טיפולים מומלץ", "type": "text", "required": False},
            {"key": "results_longevity", "label": "כמה זמן התוצאה נשמרת", "type": "text", "required": False},
            {"key": "aliases", "label": "שמות נוספים שלקוחות משתמשות בהם", "type": "list", "required": False,
             "help": "עוזר לצ׳אטבוט לזהות את הטיפול גם כשקוראות לו בשם אחר. הפרידי בפסיקים."},
        ],
    },
    "faq": {
        "table": "cb_faq",
        "pk": "faq_id",
        "label": "שאלות ותשובות",
        "description": "שאלות נפוצות והתשובות שהצ׳אטבוט עונה עליהן.",
        "order_by": "faq_id",
        "sheet": "שאלות ותשובות",
        "name_field": "canonical_question",
        "fields": [
            {"key": "category_id", "label": CATEGORY_NAME_LABEL, "type": "select", "required": False,
             "help": "אפשר לשייך לקטגוריה, או להשאיר ככללי."},
            {"key": "canonical_question", "label": "השאלה", "type": "text", "required": True},
            {"key": "answer", "label": "התשובה", "type": "textarea", "required": True,
             "help": "הצ׳אטבוט מסתמך על הנוסח הזה. מה שלא כתוב כאן — הוא לא ימציא."},
            {"key": "example_phrasings", "label": "ניסוחים נוספים של אותה שאלה", "type": "list", "required": False,
             "help": "איך עוד לקוחות שואלות את זה. הפרידי בפסיקים."},
        ],
    },
    "forward_topics": {
        "table": "cb_forward_topics",
        "pk": "topic_id",
        "label": "נושאים שמפנים לצוות",
        "description": "נושאים שהצ׳אטבוט לא עונה עליהם בעצמו אלא מפנה לטלפון.",
        "order_by": "topic_id",
        "sheet": "הפניה לצוות",
        "name_field": "topic_name",
        "fields": [
            {"key": "topic_name", "label": "הנושא", "type": "text", "required": True},
            {"key": "reason", "label": "למה מפנים לצוות", "type": "textarea", "required": False},
            {"key": "example_phrasings", "label": "ניסוחים שמפעילים את ההפניה", "type": "list", "required": False,
             "help": "הפרידי בפסיקים."},
            {"key": "phone_number", "label": "טלפון להפניה", "type": "text", "required": False},
        ],
    },
    "categories": {
        "table": "cb_categories",
        "pk": "category_id",
        "label": "קטגוריות",
        "description": "הקטגוריות הראשיות. כל טיפול משויך לאחת מהן.",
        "order_by": "category_id",
        "sheet": "קטגוריות",
        "name_field": "category_name",
        "fields": [
            {"key": "category_name", "label": "שם הקטגוריה", "type": "text", "required": True},
            {"key": "short_description", "label": "תיאור קצר", "type": "textarea", "required": False},
            {"key": "has_recommendation", "label": "שאלון התאמה פעיל", "type": "bool", "required": False,
             "help": "האם הצ׳אטבוט מציע שאלון המלצה בקטגוריה הזו."},
            {"key": "recommendation_intro", "label": "משפט פתיחה לשאלון", "type": "textarea", "required": False},
        ],
    },
}

# Sections whose rows carry a category_id and therefore get the extra readable
# category-name column in Excel.
CATEGORY_LINKED = ("treatments", "faq")

FALLBACK_PREFIXES = {"faq": "FAQ", "forward_topics": "FWD", "categories": "CAT"}


def _section(name: str) -> Dict[str, Any]:
    if name not in SECTIONS:
        raise HTTPException(status_code=404, detail=f"אין מקטע בשם {name}")
    return SECTIONS[name]


def _clean(value) -> Optional[str]:
    """Excel/JSON cell -> trimmed string, or None when empty."""
    if value is None:
        return None
    if isinstance(value, (list, tuple)):
        value = ", ".join(str(v).strip() for v in value if str(v).strip())
    s = str(value).strip()
    if s == "" or s.lower() in ("none", "nan", "null"):
        return None
    return s


def _to_bool_int(value) -> int:
    if isinstance(value, bool):
        return 1 if value else 0
    s = (_clean(value) or "").lower()
    return 1 if s in ("1", "true", "yes", "כן", "פעיל") else 0


def _field_to_db(field: Dict[str, Any], raw):
    return _to_bool_int(raw) if field["type"] == "bool" else _clean(raw)


def _db_to_display(field: Dict[str, Any], value) -> str:
    """Value as it appears in an Excel cell / a diff line."""
    if field["type"] == "bool":
        return "כן" if int(value or 0) else "לא"
    return "" if value is None else str(value)


# -- Category helpers ---------------------------------------------------------

def _categories(conn) -> List[Dict[str, Any]]:
    rows = conn.execute(
        "SELECT category_id, category_name FROM cb_categories ORDER BY category_id"
    ).fetchall()
    return [{"value": r["category_id"], "label": r["category_name"]} for r in rows]


def _category_options(conn, section_name: str) -> List[Dict[str, Any]]:
    """Dropdown options. FAQ additionally accepts the GENERAL pseudo-category."""
    options = _categories(conn)
    if section_name == "faq":
        options.append({"value": GENERAL_CATEGORY, "label": GENERAL_CATEGORY_LABEL})
    return options


def _category_name_map(conn) -> Dict[str, str]:
    m = {r["value"]: r["label"] for r in _categories(conn)}
    m[GENERAL_CATEGORY] = GENERAL_CATEGORY_LABEL
    return m


def _category_by_name(conn) -> Dict[str, str]:
    """Readable name -> code, for resolving the Excel category column."""
    m = {r["label"].strip(): r["value"] for r in _categories(conn)}
    m[GENERAL_CATEGORY_LABEL] = GENERAL_CATEGORY
    return m


def _subgroup_suggestions(conn) -> Dict[str, List[str]]:
    """Existing sub-categories per category, so the combo can suggest instead of
    forcing the owner to retype (and mistype) an existing grouping."""
    out: Dict[str, List[str]] = {}
    rows = conn.execute(
        "SELECT DISTINCT category_id, subgroup FROM cb_treatments "
        "WHERE subgroup IS NOT NULL AND subgroup <> '' ORDER BY category_id, subgroup"
    ).fetchall()
    for r in rows:
        out.setdefault(r["category_id"], []).append(r["subgroup"])
    return out


# -- Id generation ------------------------------------------------------------

def _existing_ids(conn, table: str, pk: str) -> set:
    return {r[0] for r in conn.execute(f"SELECT {pk} FROM {table}")}


def _next_id(conn, table: str, pk: str, prefix: str, taken: Optional[set] = None) -> str:
    """Next free <prefix>-NN, continuing the existing numbering."""
    ids = _existing_ids(conn, table, pk)
    if taken:
        ids |= taken
    pattern = re.compile(rf"^{re.escape(prefix)}-(\d+)$", re.IGNORECASE)
    highest = 0
    for value in ids:
        match = pattern.match(str(value or ""))
        if match:
            highest = max(highest, int(match.group(1)))
    n = highest
    while True:
        n += 1
        candidate = f"{prefix}-{n:02d}"
        if candidate not in ids:
            return candidate


def _treatment_prefix(conn, category_id: str) -> str:
    """Treatment ids are prefixed per category (CAT-03 -> CO-xx). Reuse whatever
    prefix that category already uses; a brand-new category has none, so derive
    a stable one from its number (CAT-10 -> T10-xx)."""
    rows = conn.execute(
        "SELECT treatment_id FROM cb_treatments WHERE category_id = ?", (category_id,)
    ).fetchall()
    counter: Counter = Counter()
    for r in rows:
        match = re.match(r"^([A-Za-z]+)-\d+$", str(r["treatment_id"] or ""))
        if match:
            counter[match.group(1).upper()] += 1
    if counter:
        return counter.most_common(1)[0][0]
    match = re.match(r"^CAT-(\d+)$", str(category_id or ""), re.IGNORECASE)
    return f"T{int(match.group(1)):02d}" if match else "TRT"


def _new_id_for(conn, section_name: str, data: Dict[str, Any], taken: Optional[set] = None) -> str:
    spec = SECTIONS[section_name]
    if section_name == "treatments":
        prefix = _treatment_prefix(conn, _clean(data.get("category_id")) or "")
    else:
        prefix = FALLBACK_PREFIXES[section_name]
    return _next_id(conn, spec["table"], spec["pk"], prefix, taken=taken)


# -- Validation ---------------------------------------------------------------
# The bot is hard-blocked from quoting prices and durations (SYSTEM_PROMPT rules
# 1 and 5), but text typed here is fed to it verbatim. A price written into a
# description would leak through that guard, so flag it — loudly, but without
# blocking: the owner may legitimately write "ללא עלות נוספת".
PRICE_HINT = re.compile(r"(₪|ש\"ח|שח\b|שקל|\bNIS\b|\$)", re.IGNORECASE)


def _price_warnings(spec: Dict[str, Any], data: Dict[str, Any]) -> List[str]:
    warnings = []
    for field in spec["fields"]:
        if field["type"] == "bool":
            continue
        value = _clean(data.get(field["key"]))
        if value and PRICE_HINT.search(value):
            warnings.append(
                f"בשדה \"{field['label']}\" מופיע סכום כסף. הצ׳אטבוט לא אמור למסור מחירים — "
                "מומלץ להסיר את הסכום ולהפנות לטלפון."
            )
    return warnings


def _validate(conn, section_name: str, data: Dict[str, Any], item_id: Optional[str] = None) -> List[str]:
    """Blocking errors only. Returns Hebrew messages for the admin page."""
    spec = SECTIONS[section_name]
    errors: List[str] = []

    for field in spec["fields"]:
        if field.get("required") and not _clean(data.get(field["key"])):
            errors.append(f"חסר שדה חובה: {field['label']}")

    if section_name in CATEGORY_LINKED:
        category_id = _clean(data.get("category_id"))
        if category_id:
            valid = {o["value"] for o in _category_options(conn, section_name)}
            if category_id not in valid:
                errors.append(f"הקטגוריה {category_id} לא קיימת")
        elif section_name == "treatments":
            errors.append("חובה לבחור קטגוריה לטיפול")

    if section_name == "categories":
        name = _clean(data.get("category_name"))
        if name:
            clash = conn.execute(
                "SELECT category_id FROM cb_categories "
                "WHERE category_name = ? COLLATE NOCASE AND category_id <> ?",
                (name, item_id or ""),
            ).fetchone()
            if clash:
                errors.append(f"כבר קיימת קטגוריה בשם {name} (מזהה {clash['category_id']})")

    if section_name == "treatments":
        name = _clean(data.get("treatment_name"))
        category_id = _clean(data.get("category_id"))
        if name and category_id:
            clash = conn.execute(
                "SELECT treatment_id FROM cb_treatments "
                "WHERE treatment_name = ? COLLATE NOCASE AND category_id = ? AND treatment_id <> ?",
                (name, category_id, item_id or ""),
            ).fetchone()
            if clash:
                errors.append(
                    f"כבר קיים טיפול בשם {name} בקטגוריה הזו (מזהה {clash['treatment_id']})"
                )
    return errors


# -- Read ---------------------------------------------------------------------

def _row_to_item(conn, section_name: str, row, category_names: Dict[str, str]) -> Dict[str, Any]:
    spec = SECTIONS[section_name]
    item = {"id": row[spec["pk"]]}
    for field in spec["fields"]:
        value = row[field["key"]]
        if field["type"] == "bool":
            item[field["key"]] = bool(int(value or 0))
        else:
            item[field["key"]] = "" if value is None else str(value)
    item["display_name"] = str(row[spec["name_field"]] or "")
    if section_name in CATEGORY_LINKED:
        code = row["category_id"]
        item["category_name"] = category_names.get(code, code or "")
    if section_name == "categories":
        item["treatment_count"] = conn.execute(
            "SELECT COUNT(*) FROM cb_treatments WHERE category_id = ?", (row["category_id"],)
        ).fetchone()[0]
    return item


def _list_items(conn, section_name: str) -> List[Dict[str, Any]]:
    spec = SECTIONS[section_name]
    names = _category_name_map(conn)
    rows = conn.execute(f"SELECT * FROM {spec['table']} ORDER BY {spec['order_by']}").fetchall()
    return [_row_to_item(conn, section_name, r, names) for r in rows]


def _get_row(conn, section_name: str, item_id: str):
    spec = SECTIONS[section_name]
    return conn.execute(
        f"SELECT * FROM {spec['table']} WHERE {spec['pk']} = ?", (item_id,)
    ).fetchone()


# -- Write --------------------------------------------------------------------

def _insert(conn, section_name: str, item_id: str, data: Dict[str, Any]):
    spec = SECTIONS[section_name]
    columns = [spec["pk"]] + [f["key"] for f in spec["fields"]]
    values = [item_id] + [_field_to_db(f, data.get(f["key"])) for f in spec["fields"]]
    if section_name == "treatments":
        # Bookkeeping columns the form never exposes.
        columns += ["canonical_id", "source"]
        values += [item_id, "admin"]
    placeholders = ", ".join("?" for _ in columns)
    conn.execute(
        f"INSERT INTO {spec['table']} ({', '.join(columns)}) VALUES ({placeholders})", values
    )


def _update(conn, section_name: str, item_id: str, data: Dict[str, Any]):
    """Updates only the fields in the spec — columns such as duration_min keep
    whatever the original Excel import put there."""
    spec = SECTIONS[section_name]
    assignments = ", ".join(f"{f['key']} = ?" for f in spec["fields"])
    values = [_field_to_db(f, data.get(f["key"])) for f in spec["fields"]] + [item_id]
    conn.execute(
        f"UPDATE {spec['table']} SET {assignments} WHERE {spec['pk']} = ?", values
    )


def _delete_blockers(conn, section_name: str, item_id: str) -> List[str]:
    """Reasons this row must not be deleted. Empty list = safe to delete."""
    blockers: List[str] = []
    if section_name == "categories":
        count = conn.execute(
            "SELECT COUNT(*) FROM cb_treatments WHERE category_id = ?", (item_id,)
        ).fetchone()[0]
        if count:
            blockers.append(
                f"בקטגוריה הזו יש {count} טיפולים. כדי למחוק אותה יש קודם להעביר או למחוק אותם."
            )
        faq_count = conn.execute(
            "SELECT COUNT(*) FROM cb_faq WHERE category_id = ?", (item_id,)
        ).fetchone()[0]
        if faq_count:
            blockers.append(f"יש {faq_count} שאלות ותשובות המשויכות לקטגוריה הזו.")
        quiz_count = conn.execute(
            "SELECT COUNT(*) FROM cb_questions WHERE category_id = ?", (item_id,)
        ).fetchone()[0]
        if quiz_count:
            blockers.append("לקטגוריה הזו מוגדר שאלון התאמה. יש להסיר אותו קודם מקובץ האקסל.")
    if section_name == "treatments":
        terminal = conn.execute(
            "SELECT COUNT(*) FROM cb_questions WHERE terminal_treatment_id = ?", (item_id,)
        ).fetchone()[0]
        if terminal:
            blockers.append(
                "הטיפול הזה הוא תוצאה ישירה של שאלה בשאלון ההתאמה. "
                "יש להסיר אותו מהשאלון לפני המחיקה."
            )
    return blockers


def _cascade_delete(conn, section_name: str, item_id: str) -> Dict[str, int]:
    """Removes rows that would otherwise point at a treatment that no longer
    exists — a dangling score would make the quiz recommend a missing treatment."""
    removed = {}
    if section_name == "treatments":
        cur = conn.execute("DELETE FROM cb_scoring WHERE treatment_id = ?", (item_id,))
        if cur.rowcount:
            removed["quiz_scores"] = cur.rowcount
    return removed


# -- Backups ------------------------------------------------------------------

def _backup_db(reason: str) -> str:
    """Consistent snapshot of chatbot.db before anything bulk touches it."""
    BACKUP_DIR.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    destination = BACKUP_DIR / f"chatbot-{stamp}-{reason}.db"
    source = sqlite3.connect(str(CHATBOT_DB_PATH))
    target = sqlite3.connect(str(destination))
    try:
        with target:
            source.backup(target)
    finally:
        source.close()
        target.close()
    snapshots = sorted(BACKUP_DIR.glob("chatbot-*.db"))
    for stale in snapshots[:-KEEP_BACKUPS]:
        try:
            stale.unlink()
        except OSError:
            pass
    return destination.name


# -- Excel columns ------------------------------------------------------------

def _export_columns(section_name: str) -> List[Dict[str, Any]]:
    """One entry per spreadsheet column, in order.

    kind 'id'            -> the row's symbol (CO-05). Blank means "add this row".
    kind 'category_name' -> the readable category, which is what she edits.
    kind 'category_code' -> the matching symbol, kept alongside so every code in
                            the sheet can be traced back to a name.
    """
    spec = SECTIONS[section_name]
    columns: List[Dict[str, Any]] = [{"kind": "id", "header": ID_COLUMN_LABEL}]
    for field in spec["fields"]:
        if field["key"] == "category_id":
            columns.append({"kind": "category_name", "header": CATEGORY_NAME_LABEL, "field": field})
            columns.append({"kind": "category_code", "header": CATEGORY_CODE_LABEL, "field": field})
        else:
            columns.append({"kind": "field", "header": field["label"], "field": field})
    return columns


def _build_workbook(conn, section_name: str):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    spec = SECTIONS[section_name]
    columns = _export_columns(section_name)
    names = _category_name_map(conn)
    rows = conn.execute(f"SELECT * FROM {spec['table']} ORDER BY {spec['order_by']}").fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = spec["sheet"]
    ws.sheet_view.rightToLeft = True

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="C4795A")
    id_fill = PatternFill("solid", fgColor="F0E4DA")

    for index, column in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=index, value=column["header"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

    for row_index, row in enumerate(rows, start=2):
        for column_index, column in enumerate(columns, start=1):
            if column["kind"] == "id":
                value = row[spec["pk"]]
            elif column["kind"] == "category_name":
                value = names.get(row["category_id"], "")
            elif column["kind"] == "category_code":
                value = row["category_id"] or ""
            else:
                value = _db_to_display(column["field"], row[column["field"]["key"]])
            cell = ws.cell(row=row_index, column=column_index, value=value)
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
            if column["kind"] in ("id", "category_code"):
                cell.fill = id_fill

    for index, column in enumerate(columns, start=1):
        letter = get_column_letter(index)
        if column["kind"] in ("id", "category_code"):
            ws.column_dimensions[letter].width = 16
        elif column["kind"] == "category_name":
            ws.column_dimensions[letter].width = 22
        elif column["field"]["type"] in ("textarea", "list"):
            ws.column_dimensions[letter].width = 48
        else:
            ws.column_dimensions[letter].width = 24
    ws.freeze_panes = "A2"

    _add_guide_sheet(wb, conn, section_name)
    return wb


def _add_guide_sheet(wb, conn, section_name: str):
    """Explains what the codes in the sheet mean. The workbook is full of symbols
    (CAT-03, CO-05); without this the owner has no way to read them."""
    from openpyxl.styles import Alignment, Font

    spec = SECTIONS[section_name]
    ws = wb.create_sheet("מקרא והוראות")
    ws.sheet_view.rightToLeft = True
    title = Font(bold=True, size=13)
    heading = Font(bold=True)

    lines = [
        (f"{spec['label']} — איך לערוך את הקובץ", title),
        ("", None),
        ("1. עריכת שורה קיימת: שני את הטקסט בעמודה שרצית. אל תשני את עמודת \"מזהה\".", None),
        ("2. הוספת שורה חדשה: הוסיפי שורה בסוף והשאירי את עמודת \"מזהה\" ריקה. המערכת תיצור מזהה לבד.", None),
        ("3. מחיקת שורה: מחקי את כל השורה מהקובץ.", None),
        ("4. שמירה: שמרי את הקובץ והעלי אותו בעמוד הניהול. יוצג לך סיכום השינויים לאישור לפני שמשהו נשמר.", None),
        ("", None),
        ("עמודות שמסומנות בצבע הן לעיון בלבד ואין צורך לערוך אותן.", None),
    ]
    if section_name in CATEGORY_LINKED:
        lines.append(
            ("בעמודת \"קטגוריה\" יש לכתוב שם קטגוריה מהרשימה שלמטה, בדיוק כפי שהוא מופיע שם.", None)
        )
    for row_index, (text, font) in enumerate(lines, start=1):
        cell = ws.cell(row=row_index, column=1, value=text)
        if font:
            cell.font = font
        cell.alignment = Alignment(horizontal="right")

    if section_name in CATEGORY_LINKED or section_name == "categories":
        ws.cell(row=ws.max_row + 2, column=1, value="הקטגוריות הקיימות").font = heading
        header_row = ws.max_row + 1
        for index, label in enumerate(["שם הקטגוריה", "קוד", "כמה טיפולים יש בה", "קידומת המזהה של הטיפולים"], start=1):
            ws.cell(row=header_row, column=index, value=label).font = heading
        for r in conn.execute(
            "SELECT category_id, category_name FROM cb_categories ORDER BY category_id"
        ).fetchall():
            count = conn.execute(
                "SELECT COUNT(*) FROM cb_treatments WHERE category_id = ?", (r["category_id"],)
            ).fetchone()[0]
            row_index = ws.max_row + 1
            ws.cell(row=row_index, column=1, value=r["category_name"])
            ws.cell(row=row_index, column=2, value=r["category_id"])
            ws.cell(row=row_index, column=3, value=count)
            ws.cell(row=row_index, column=4, value=f"{_treatment_prefix(conn, r['category_id'])}-")
        if section_name == "faq":
            row_index = ws.max_row + 1
            ws.cell(row=row_index, column=1, value=GENERAL_CATEGORY_LABEL)
            ws.cell(row=row_index, column=2, value=GENERAL_CATEGORY)

    for index, width in enumerate([34, 14, 20, 26], start=1):
        ws.column_dimensions[chr(64 + index)].width = width
    return ws


# -- Import: parse + diff -----------------------------------------------------

def _parse_workbook(conn, section_name: str, payload: bytes) -> Dict[str, Any]:
    """Sheet -> desired rows, with per-row errors. Reads nothing into the DB."""
    from openpyxl import load_workbook

    spec = SECTIONS[section_name]
    columns = _export_columns(section_name)
    by_name = _category_by_name(conn)
    valid_codes = {option["value"] for option in _category_options(conn, section_name)}

    try:
        wb = load_workbook(io.BytesIO(payload), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="הקובץ אינו קובץ אקסל תקין (xlsx)")

    ws = wb[spec["sheet"]] if spec["sheet"] in wb.sheetnames else wb.worksheets[0]
    sheet_rows = list(ws.iter_rows(values_only=True))
    if not sheet_rows:
        raise HTTPException(status_code=400, detail="הגיליון ריק")

    headers = [(_clean(h) or "") for h in sheet_rows[0]]
    header_index = {}
    for position, header in enumerate(headers):
        if header and header not in header_index:
            header_index[header] = position

    missing = [
        column["header"] for column in columns
        if column["kind"] == "field" and column["field"].get("required")
        and column["header"] not in header_index
    ]
    if section_name == "treatments" and CATEGORY_NAME_LABEL not in header_index \
            and CATEGORY_CODE_LABEL not in header_index:
        missing.append(CATEGORY_NAME_LABEL)
    if ID_COLUMN_LABEL not in header_index:
        missing.append(ID_COLUMN_LABEL)
    if missing:
        raise HTTPException(
            status_code=400,
            detail="חסרות עמודות בקובץ: " + ", ".join(dict.fromkeys(missing))
            + ". מומלץ להוריד את הקובץ מחדש ולערוך אותו.",
        )

    def cell(row, header):
        position = header_index.get(header)
        return row[position] if position is not None and position < len(row) else None

    parsed: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    seen_ids: Dict[str, int] = {}

    for offset, row in enumerate(sheet_rows[1:], start=2):
        if all(_clean(value) is None for value in row):
            continue
        item_id = _clean(cell(row, ID_COLUMN_LABEL))
        data: Dict[str, Any] = {}
        row_errors: List[str] = []

        for column in columns:
            if column["kind"] in ("id", "category_code"):
                continue
            field = column["field"]
            if column["kind"] == "category_name":
                name = _clean(cell(row, CATEGORY_NAME_LABEL))
                code = _clean(cell(row, CATEGORY_CODE_LABEL))
                resolved = by_name.get(name) if name else None
                if resolved is None and name:
                    row_errors.append(f"אין קטגוריה בשם \"{name}\"")
                if resolved is None and code:
                    resolved = code if code in valid_codes else None
                    if resolved is None:
                        row_errors.append(f"קוד הקטגוריה \"{code}\" לא קיים")
                data["category_id"] = resolved
            else:
                data[field["key"]] = _field_to_db(field, cell(row, column["header"]))

        if item_id:
            if item_id in seen_ids:
                row_errors.append(f"המזהה {item_id} מופיע פעמיים בקובץ (גם בשורה {seen_ids[item_id]})")
            else:
                seen_ids[item_id] = offset
            if not _get_row(conn, section_name, item_id):
                row_errors.append(
                    f"המזהה {item_id} לא קיים במערכת. כדי להוסיף שורה חדשה יש להשאיר את תא המזהה ריק."
                )
        for field in spec["fields"]:
            if field.get("required") and field["key"] != "category_id" and not _clean(data.get(field["key"])):
                row_errors.append(f"חסר ערך בעמודה {field['label']}")

        if row_errors:
            errors.extend({"row": offset, "message": message} for message in row_errors)
        else:
            parsed.append({"row": offset, "id": item_id, "data": data})

    return {"rows": parsed, "errors": errors}


def _diff(conn, section_name: str, parsed_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Desired sheet state vs. what is in the DB right now."""
    spec = SECTIONS[section_name]
    names = _category_name_map(conn)
    existing = {
        r[spec["pk"]]: r
        for r in conn.execute(f"SELECT * FROM {spec['table']}").fetchall()
    }

    changes: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    kept: set = set()
    counts = {"new": 0, "updated": 0, "removed": 0, "unchanged": 0}

    def label_for(data: Dict[str, Any]) -> str:
        return _clean(data.get(spec["name_field"])) or "(ללא שם)"

    for entry in parsed_rows:
        item_id, data = entry["id"], entry["data"]
        if not item_id:
            counts["new"] += 1
            changes.append({
                "kind": "new",
                "id": "",
                "row": entry["row"],
                "name": label_for(data),
                "category": names.get(data.get("category_id"), ""),
                "fields": [],
            })
            continue

        kept.add(item_id)
        current = existing[item_id]
        differing = []
        for field in spec["fields"]:
            before = _db_to_display(field, current[field["key"]])
            after = _db_to_display(field, data.get(field["key"]))
            if (before or "").strip() != (after or "").strip():
                differing.append({
                    "label": field["label"],
                    "before": names.get(before, before) if field["key"] == "category_id" else before,
                    "after": names.get(after, after) if field["key"] == "category_id" else after,
                })
        if differing:
            counts["updated"] += 1
            changes.append({
                "kind": "updated",
                "id": item_id,
                "row": entry["row"],
                "name": label_for(data),
                "category": names.get(data.get("category_id"), ""),
                "fields": differing,
            })
        else:
            counts["unchanged"] += 1

    for item_id, current in existing.items():
        if item_id in kept:
            continue
        counts["removed"] += 1
        blockers = _delete_blockers(conn, section_name, item_id)
        changes.append({
            "kind": "removed",
            "id": item_id,
            "row": None,
            "name": str(current[spec["name_field"]] or ""),
            "category": names.get(current["category_id"], "") if section_name in CATEGORY_LINKED else "",
            "fields": [],
            "blockers": blockers,
        })
        for blocker in blockers:
            errors.append({"row": None, "message": f"לא ניתן למחוק את {item_id}: {blocker}"})

    warnings: List[str] = []
    total_existing = len(existing)
    if counts["removed"] and total_existing and counts["removed"] / total_existing > 0.3:
        warnings.append(
            f"שימי לב: הקובץ מוחק {counts['removed']} מתוך {total_existing} שורות. "
            "אם לא התכוונת למחוק אותן — בטלי, הורידי את הקובץ מחדש וערכי אותו."
        )
    if not parsed_rows:
        warnings.append("הקובץ לא מכיל אף שורת מידע.")

    return {"summary": counts, "changes": changes, "errors": errors, "warnings": warnings}


# -- Import staging -----------------------------------------------------------

def _stage(section_name: str, parsed_rows: List[Dict[str, Any]], actor_id: Optional[int]) -> str:
    STAGING_DIR.mkdir(exist_ok=True)
    _prune_staging()
    token = secrets.token_urlsafe(16)
    (STAGING_DIR / f"{token}.json").write_text(
        json.dumps({
            "section": section_name,
            "actor_id": actor_id,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "rows": parsed_rows,
        }, ensure_ascii=False),
        encoding="utf-8",
    )
    return token


def _load_stage(token: str, section_name: str, actor_id: Optional[int]) -> List[Dict[str, Any]]:
    path = STAGING_DIR / f"{re.sub(r'[^A-Za-z0-9_-]', '', token)}.json"
    if not path.exists():
        raise HTTPException(status_code=400, detail="תוקף האישור פג. יש להעלות את הקובץ מחדש.")
    payload = json.loads(path.read_text(encoding="utf-8"))
    created = datetime.fromisoformat(payload["created_at"])
    if datetime.now(timezone.utc) - created > timedelta(minutes=STAGING_TTL_MINUTES):
        path.unlink(missing_ok=True)
        raise HTTPException(status_code=400, detail="תוקף האישור פג. יש להעלות את הקובץ מחדש.")
    if payload["section"] != section_name or payload.get("actor_id") != actor_id:
        raise HTTPException(status_code=400, detail="האישור אינו תואם לקובץ שהועלה.")
    path.unlink(missing_ok=True)
    return payload["rows"]


def _prune_staging():
    cutoff = datetime.now(timezone.utc) - timedelta(minutes=STAGING_TTL_MINUTES)
    for path in STAGING_DIR.glob("*.json"):
        try:
            created = datetime.fromisoformat(json.loads(path.read_text(encoding="utf-8"))["created_at"])
            if created < cutoff:
                path.unlink(missing_ok=True)
        except Exception:
            path.unlink(missing_ok=True)


# -- Quiz (read-only) ---------------------------------------------------------

def _quiz_overview(conn) -> List[Dict[str, Any]]:
    """The recommendation quiz stays read-only here: it is a weighted matrix
    (answer X gives treatment Y N points) and editing it safely needs its own
    screen. Shown so the owner can see it exists and what it asks."""
    names = _category_name_map(conn)
    treatment_names = {
        r["treatment_id"]: r["treatment_name"]
        for r in conn.execute("SELECT treatment_id, treatment_name FROM cb_treatments")
    }
    out: List[Dict[str, Any]] = []
    category_ids = [
        r[0] for r in conn.execute(
            "SELECT DISTINCT category_id FROM cb_questions ORDER BY category_id"
        )
    ]
    for category_id in category_ids:
        questions: Dict[str, Dict[str, Any]] = {}
        rows = conn.execute(
            "SELECT * FROM cb_questions WHERE category_id = ? "
            "ORDER BY question_order, option_order",
            (category_id,),
        ).fetchall()
        for r in rows:
            question = questions.setdefault(r["question_id"], {
                "question_id": r["question_id"],
                "text": r["question_text"],
                "options": [],
            })
            question["options"].append({
                "label": r["option_label"],
                "leads_to": treatment_names.get(r["terminal_treatment_id"], "") if r["terminal_treatment_id"] else "",
            })
        score_count = conn.execute(
            "SELECT COUNT(*) FROM cb_scoring WHERE category_id = ?", (category_id,)
        ).fetchone()[0]
        out.append({
            "category_id": category_id,
            "category_name": names.get(category_id, category_id),
            "questions": list(questions.values()),
            "score_rows": score_count,
        })
    return out


# -- Router -------------------------------------------------------------------

def build_router(require_admin, log_audit) -> APIRouter:
    """main.py owns auth and the audit log; they are injected so this module
    never imports main (which imports this one)."""
    router = APIRouter(prefix="/admin/ai", tags=["ai-content"])

    def _audit(actor, action: str, section_name: str, item_id: str, details: str = ""):
        log_audit(
            action=action,
            actor=actor,
            target_type=f"chatbot_{section_name}",
            target_username=item_id,
            details=details,
        )

    @router.get("/overview")
    def overview(_: dict = Depends(require_admin)):
        conn = get_chatbot_db()
        try:
            sections = []
            for name, spec in SECTIONS.items():
                count = conn.execute(f"SELECT COUNT(*) FROM {spec['table']}").fetchone()[0]
                sections.append({
                    "key": name,
                    "label": spec["label"],
                    "description": spec["description"],
                    "count": count,
                })
            categories = []
            for r in conn.execute(
                "SELECT category_id, category_name FROM cb_categories ORDER BY category_id"
            ).fetchall():
                categories.append({
                    "category_id": r["category_id"],
                    "category_name": r["category_name"],
                    "treatment_count": conn.execute(
                        "SELECT COUNT(*) FROM cb_treatments WHERE category_id = ?",
                        (r["category_id"],),
                    ).fetchone()[0],
                    "prefix": _treatment_prefix(conn, r["category_id"]),
                })
            return {"ok": True, "sections": sections, "categories": categories}
        finally:
            conn.close()

    @router.get("/quiz")
    def quiz(_: dict = Depends(require_admin)):
        conn = get_chatbot_db()
        try:
            return {"ok": True, "categories": _quiz_overview(conn)}
        finally:
            conn.close()

    @router.get("/{section_name}/export")
    def export_section(section_name: str, _: dict = Depends(require_admin)):
        spec = _section(section_name)
        conn = get_chatbot_db()
        try:
            wb = _build_workbook(conn, section_name)
        finally:
            conn.close()
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        ascii_name = f"meday-{section_name}.xlsx"
        pretty = f"MeDay - {spec['label']}.xlsx"
        quoted = "".join(
            character if character.isalnum() or character in "-_. " else f"%{ord(character):02X}"
            for character in pretty.encode("utf-8").decode("latin-1", "ignore")
        )
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{quoted}"},
        )

    @router.post("/{section_name}/import/preview")
    async def import_preview(
        section_name: str,
        file: UploadFile = File(...),
        actor: dict = Depends(require_admin),
    ):
        _section(section_name)
        payload = await file.read()
        if not payload:
            raise HTTPException(status_code=400, detail="הקובץ ריק")
        if len(payload) > MAX_UPLOAD_BYTES:
            raise HTTPException(status_code=400, detail="הקובץ גדול מדי (מקסימום 5MB)")
        conn = get_chatbot_db()
        try:
            parsed = _parse_workbook(conn, section_name, payload)
            if parsed["errors"]:
                # Rows that failed to parse are not in the desired state, so a diff
                # would report them as deletions. Ask her to fix the file first.
                return {
                    "ok": False,
                    "blocked": True,
                    "summary": {"new": 0, "updated": 0, "removed": 0, "unchanged": 0},
                    "changes": [],
                    "errors": parsed["errors"][:100],
                    "warnings": ["יש לתקן את השגיאות בקובץ ולהעלות אותו שוב. שום דבר לא נשמר."],
                }
            result = _diff(conn, section_name, parsed["rows"])
            blocked = bool(result["errors"])
            token = None if blocked else _stage(section_name, parsed["rows"], actor.get("id"))
            return {
                "ok": not blocked,
                "blocked": blocked,
                "token": token,
                "summary": result["summary"],
                "changes": result["changes"][:200],
                "changes_truncated": len(result["changes"]) > 200,
                "errors": result["errors"][:100],
                "warnings": result["warnings"],
            }
        finally:
            conn.close()

    @router.post("/{section_name}/import/apply")
    def import_apply(
        section_name: str,
        body: Dict[str, Any],
        actor: dict = Depends(require_admin),
    ):
        spec = _section(section_name)
        token = _clean(body.get("token"))
        if not token:
            raise HTTPException(status_code=400, detail="חסר אישור לייבוא")
        rows = _load_stage(token, section_name, actor.get("id"))

        conn = get_chatbot_db()
        try:
            result = _diff(conn, section_name, rows)
            if result["errors"]:
                raise HTTPException(status_code=400, detail=result["errors"][0]["message"])

            backup_name = _backup_db(f"import-{section_name}")
            kept = {row["id"] for row in rows if row["id"]}
            existing_ids = _existing_ids(conn, spec["table"], spec["pk"])
            created, updated, removed = 0, 0, 0
            cascaded = 0

            conn.execute("BEGIN")
            try:
                minted: set = set()
                for row in rows:
                    if row["id"]:
                        _update(conn, section_name, row["id"], row["data"])
                        updated += 1
                    else:
                        new_id = _new_id_for(conn, section_name, row["data"], taken=minted)
                        minted.add(new_id)
                        _insert(conn, section_name, new_id, row["data"])
                        created += 1
                for item_id in existing_ids - kept:
                    cascaded += sum(_cascade_delete(conn, section_name, item_id).values())
                    conn.execute(
                        f"DELETE FROM {spec['table']} WHERE {spec['pk']} = ?", (item_id,)
                    )
                    removed += 1
                conn.commit()
            except Exception:
                conn.rollback()
                raise
        finally:
            conn.close()

        _audit(
            actor, "chatbot_content_import", section_name, "-",
            f"נוספו {created}, עודכנו {updated}, נמחקו {removed} (גיבוי: {backup_name})",
        )
        return {
            "ok": True,
            "created": created,
            "updated": updated,
            "removed": removed,
            "cascaded_quiz_rows": cascaded,
            "backup": backup_name,
        }

    @router.get("/{section_name}")
    def list_section(section_name: str, _: dict = Depends(require_admin)):
        spec = _section(section_name)
        conn = get_chatbot_db()
        try:
            return {
                "ok": True,
                "key": section_name,
                "label": spec["label"],
                "description": spec["description"],
                "fields": spec["fields"],
                "category_options": _category_options(conn, section_name),
                "subgroups": _subgroup_suggestions(conn) if section_name == "treatments" else {},
                "items": _list_items(conn, section_name),
            }
        finally:
            conn.close()

    @router.post("/{section_name}")
    def create_item(section_name: str, body: Dict[str, Any], actor: dict = Depends(require_admin)):
        spec = _section(section_name)
        conn = get_chatbot_db()
        try:
            errors = _validate(conn, section_name, body)
            if errors:
                raise HTTPException(status_code=400, detail=errors[0])
            item_id = _new_id_for(conn, section_name, body)
            conn.execute("BEGIN")
            try:
                _insert(conn, section_name, item_id, body)
                conn.commit()
            except Exception:
                conn.rollback()
                raise
            names = _category_name_map(conn)
            item = _row_to_item(conn, section_name, _get_row(conn, section_name, item_id), names)
            warnings = _price_warnings(spec, body)
        finally:
            conn.close()
        _audit(actor, "chatbot_content_create", section_name, item_id, item.get("display_name", ""))
        return {"ok": True, "item": item, "warnings": warnings}

    @router.put("/{section_name}/{item_id}")
    def update_item(
        section_name: str, item_id: str, body: Dict[str, Any],
        actor: dict = Depends(require_admin),
    ):
        spec = _section(section_name)
        conn = get_chatbot_db()
        try:
            if not _get_row(conn, section_name, item_id):
                raise HTTPException(status_code=404, detail="הפריט לא נמצא")
            errors = _validate(conn, section_name, body, item_id=item_id)
            if errors:
                raise HTTPException(status_code=400, detail=errors[0])
            conn.execute("BEGIN")
            try:
                _update(conn, section_name, item_id, body)
                conn.commit()
            except Exception:
                conn.rollback()
                raise
            names = _category_name_map(conn)
            item = _row_to_item(conn, section_name, _get_row(conn, section_name, item_id), names)
            warnings = _price_warnings(spec, body)
        finally:
            conn.close()
        _audit(actor, "chatbot_content_update", section_name, item_id, item.get("display_name", ""))
        return {"ok": True, "item": item, "warnings": warnings}

    @router.delete("/{section_name}/{item_id}")
    def delete_item(section_name: str, item_id: str, actor: dict = Depends(require_admin)):
        spec = _section(section_name)
        conn = get_chatbot_db()
        try:
            row = _get_row(conn, section_name, item_id)
            if not row:
                raise HTTPException(status_code=404, detail="הפריט לא נמצא")
            blockers = _delete_blockers(conn, section_name, item_id)
            if blockers:
                raise HTTPException(status_code=409, detail=blockers[0])
            name = str(row[spec["name_field"]] or "")
            conn.execute("BEGIN")
            try:
                cascaded = _cascade_delete(conn, section_name, item_id)
                conn.execute(f"DELETE FROM {spec['table']} WHERE {spec['pk']} = ?", (item_id,))
                conn.commit()
            except Exception:
                conn.rollback()
                raise
        finally:
            conn.close()
        _audit(actor, "chatbot_content_delete", section_name, item_id, name)
        return {"ok": True, "cascaded": cascaded}

    return router
